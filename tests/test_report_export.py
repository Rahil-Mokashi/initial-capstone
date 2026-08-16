import csv
from datetime import datetime, timezone
from decimal import Decimal
from types import SimpleNamespace

import openpyxl
import pytest
from pypdf import PdfReader

from app.services.report_export import (
    build_sale_receipt_html,
    build_table_report_html,
    export_fuel_summary_excel,
    export_fuel_summary_pdf,
    export_sale_receipt_pdf,
    export_table_csv,
    export_table_excel,
    export_table_pdf,
)
from app.services.report_service import FuelTypeSummary, TableReport


@pytest.fixture()
def summaries():
    return [
        FuelTypeSummary(
            fuel_type="Petrol",
            fuel_id="f1",
            tank_count=2,
            total_capacity=Decimal("20000.000"),
            total_current_stock=Decimal("15000.500"),
            nozzle_count=4,
            active_nozzle_count=3,
            latest_variance_percent=Decimal("0.750"),
            latest_variance_classification="normal",
        ),
        FuelTypeSummary(
            fuel_type="Diesel",
            fuel_id="f2",
            tank_count=1,
            total_capacity=Decimal("10000.000"),
            total_current_stock=Decimal("4000.000"),
            nozzle_count=2,
            active_nozzle_count=2,
            latest_variance_percent=None,
            latest_variance_classification=None,
        ),
    ]


def test_export_pdf_creates_a_readable_file(tmp_path, summaries):
    pdf_path = str(tmp_path / "fuel_summary.pdf")
    export_fuel_summary_pdf(summaries, pdf_path)

    reader = PdfReader(pdf_path)
    assert len(reader.pages) >= 1
    text = reader.pages[0].extract_text()
    assert "Petrol" in text
    assert "Diesel" in text


def test_export_excel_creates_expected_rows(tmp_path, summaries):
    xlsx_path = str(tmp_path / "fuel_summary.xlsx")
    export_fuel_summary_excel(summaries, xlsx_path)

    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active

    assert sheet["A1"].value == "Fuel Type"
    assert sheet["A2"].value == "Petrol"
    assert sheet["A3"].value == "Diesel"
    assert sheet.max_row == 3


def test_export_excel_handles_missing_variance(tmp_path, summaries):
    xlsx_path = str(tmp_path / "fuel_summary.xlsx")
    export_fuel_summary_excel(summaries, xlsx_path)

    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active

    diesel_row = sheet[3]
    variance_cell = diesel_row[6]
    assert variance_cell.value is None


def test_export_pdf_handles_empty_summary_list(tmp_path):
    pdf_path = str(tmp_path / "empty.pdf")
    export_fuel_summary_pdf([], pdf_path)

    reader = PdfReader(pdf_path)
    assert len(reader.pages) >= 1


def test_export_excel_handles_empty_summary_list(tmp_path):
    xlsx_path = str(tmp_path / "empty.xlsx")
    export_fuel_summary_excel([], xlsx_path)

    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active
    assert sheet.max_row == 1


@pytest.fixture()
def table_report():
    return TableReport(
        title="Sales Report",
        headers=["Fuel Type", "Sales", "Amount"],
        rows=[["Petrol", "3", "1000.00"], ["Total", "3", "1000.00"]],
    )


def test_export_table_pdf_creates_a_readable_file(tmp_path, table_report):
    pdf_path = str(tmp_path / "sales.pdf")
    export_table_pdf(table_report, pdf_path)

    reader = PdfReader(pdf_path)
    text = reader.pages[0].extract_text()
    assert "Sales Report" in text
    assert "Petrol" in text


def test_export_table_excel_creates_expected_rows(tmp_path, table_report):
    xlsx_path = str(tmp_path / "sales.xlsx")
    export_table_excel(table_report, xlsx_path)

    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active
    assert sheet["A1"].value == "Fuel Type"
    assert sheet["A2"].value == "Petrol"
    assert sheet.max_row == 3


def test_export_table_csv_creates_expected_rows(tmp_path, table_report):
    csv_path = str(tmp_path / "sales.csv")
    export_table_csv(table_report, csv_path)

    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    assert rows[0] == ["Fuel Type", "Sales", "Amount"]
    assert rows[1] == ["Petrol", "3", "1000.00"]
    assert rows[2] == ["Total", "3", "1000.00"]


def test_build_table_report_html_includes_title_and_rows():
    report = TableReport(title="Expense Summary Report", headers=["Category", "Approved"], rows=[["Electricity", "300.00"]])
    html = build_table_report_html(report)
    assert "Expense Summary Report" in html
    assert "Electricity" in html
    assert "300.00" in html


@pytest.fixture()
def sale_stub():
    return SimpleNamespace(
        receipt_number="RCPT-000001",
        sale_at=datetime(2026, 8, 16, 10, 30, tzinfo=timezone.utc),
        fuel=SimpleNamespace(fuel_type="Petrol"),
        nozzle=SimpleNamespace(code="N1"),
        quantity=Decimal("10.000"),
        rate_per_liter=Decimal("100.00"),
        amount=Decimal("1000.00"),
        payment_method="cash",
        customer=None,
    )


@pytest.fixture()
def payment_stub():
    return SimpleNamespace(reference_number=None)


def test_export_sale_receipt_pdf_creates_a_readable_file(tmp_path, sale_stub, payment_stub):
    pdf_path = str(tmp_path / "receipt.pdf")
    export_sale_receipt_pdf(sale_stub, payment_stub, pdf_path)

    reader = PdfReader(pdf_path)
    text = reader.pages[0].extract_text()
    assert "RCPT-000001" in text
    assert "Petrol" in text
    assert "1000.00" in text


def test_build_sale_receipt_html_includes_customer_when_present(sale_stub, payment_stub):
    sale_stub.customer = SimpleNamespace(name="Ravi Transports")
    html = build_sale_receipt_html(sale_stub, payment_stub)
    assert "RCPT-000001" in html
    assert "Ravi Transports" in html


def test_build_sale_receipt_html_includes_reference_when_present(sale_stub):
    payment = SimpleNamespace(reference_number="UPI-REF-999")
    html = build_sale_receipt_html(sale_stub, payment)
    assert "UPI-REF-999" in html
